import os
import numpy as np
import torch
import openpyxl



def createExcel(save_path = '../results/metrics.xlsx'):
    data_name = ['loss', 'accuracy', 'precision', 'recall', 'f1', 'DSC', 'HM', 'IOU', 'val_loss', 'val_accuracy',
                 'val_precision', 'val_recall', 'val_f1', 'val_DSC', 'val_HM', 'val_IOU']
    mWorkBook = openpyxl.Workbook()

    mSheet = mWorkBook.active

    mSheet.cell(row=1, column=1).value = 'epoch'
    for i in range(16):
        mSheet.cell(row=1, column=i+2).value = data_name[i]

    mWorkBook.save(save_path)
    mWorkBook.close()


#the first column is epoch, and the other 16 columns is scores
#train metrics：'loss', 'accuracy', 'precision', 'recall', 'f1', 'DSC', 'HM', 'IOU'
# val metrics：'val_loss', 'val_accuracy', 'val_precision', 'val_recall', 'val_f1', 'val_DSC', 'val_HM', 'val_IOU'
def to_Excel(EXCEL_PATH, metrics, epoch):
    if os.path.exists(EXCEL_PATH) == False:
        createExcel()
    mWorkBook = openpyxl.load_workbook(EXCEL_PATH)
    mSheet = mWorkBook.active
    row = epoch + 1
    
    mSheet.cell(row=row, column=1).value = epoch

    for i in range(16):
        mSheet.cell(row=row, column=i+2).value = metrics[i]
    # 保存工作簿
    mWorkBook.save(EXCEL_PATH)
    # 关闭
    mWorkBook.close()



if __name__ == '__main__':
    createExcel()
    to_Excel('../results/metrics.xlsx', np.zeros(shape=16), 1)